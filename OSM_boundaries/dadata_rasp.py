from dadata import Dadata
from clickhouse_connect import get_client
import pandas as pd
from openpyxl import load_workbook


# dadata api
token = "0b98b6b33ac216cb85afafb8126e2b84726dd012"
secret = "101e666124845d7cbd05889746cc718f15fed1df"
dadata = Dadata(token, secret)

# clickhouse
client = get_client(host='clickhouse.moscow', username='GrushkoIV', port = '8123', password='jNbrvzd1IcF0Yx5I', database='grushko_iv')

with client:
    df = client.query_df(f'''
            SELECT * FROM MdlpAdresCorrectedDadata              
            ''') 
    # print(df.dtypes)
rows = []    
for _, r in df[15000:].iterrows():
    addr_str = r['address_address_description']
    id = r['id']
    result = dadata.clean("address", addr_str)
    if (result['geo_lat'] is not None and result['geo_lon'] is not None):
        lat = result['geo_lat']
        lon = result['geo_lon']
        accurace = result['qc_geo']
        # print(lat, lon)
        df['lat'] = lat
        df['lon'] = lon
        df['accurace'] = accurace 
        rows.append({"id": id, "lat": lat, "lon": lon, "accurace": accurace})

df_coord = pd.DataFrame(rows)
# wb = load_workbook("coords.xlsx")   # файл уже существует
# ws = wb.active                      # возьмём первый лист
# for row in rows:
#     ws.append(row)

# wb.save("coords.xlsx")    

# pd.DataFrame(rows, columns=["id","lat","lon","accurace"]).to_excel("coords.xlsx", index=False)
book = load_workbook("coords.xlsx")
startrow = book['Sheet1'].max_row if 'Sheet1' in book.sheetnames else 0
with pd.ExcelWriter("coords.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
    df_coord.to_excel(
        writer,
        index=False,
        header=(startrow == 0),     # шапку пишем только если лист новый/пустой
        startrow=startrow,
        sheet_name='Sheet1',
        )
# excel = pd.DataFrame(result_records, columns=['id', 'federal_subject', 'boundary_name'])

# with pd.ExcelWriter(out_path, engine='openpyxl') as xw:
#     excel.to_excel(xw, index=False, sheet_name='results')

#     print(f'Excel сохранён: {out_path}')


#         client.command(f'''
#             ALTER TABLE MdlpAdresCorrectedDadata UPDATE lat = CAST({lat}  AS Nullable(Float64)), lon = CAST({lon}  AS Nullable(Float64)) WHERE id = {id}   
#             ''')
#         print(f'По МД {id} координаты обновлены')


# df = df.copy()  # чтобы спокойно изменять
# for i, row in df.iterrows():
    